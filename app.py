# Biblioteca usada para acessar planilhas
import openpyxl
# Biblioteca usada para acessar sites
from selenium import webdriver
# Encontrar elementos dento das páginas e interagir com eles
from selenium.webdriver.common.by import By
# Biblioteca usada para pausar
from time import sleep

# 1 - Entrar na planilha e extrair o CPF do cliente.
planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
# Seleciona a página dos dados
pagina_clientes = planilha_clientes['Sheet1']

# Acessar o site pelo chrome
driver = webdriver.Chrome()
# O site que deve ser acessado
driver.get('https://consultcpf-devaprender.netlify.app/')

# Permite ler cada linha de uma planilha. Neste caso começamos pela linha 2 por causa do 'min_row=2'.
# O 'values_only = true' significa que só vai retornar os valores verdadeiros.
for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
    # Extrai as informações
    nome, valor, cpf, vencimento = linha
    # Printa os valores
    # print(nome)
    # print(valor)
    # print(cpf)
    # print(vencimento)

    # 2 - Entrar no site https://consultcpf-devaprender.netlify.app/ e usar o cpf da planilha para pesquisar o status do pagamento daquele cliente.
    # Dar uma pausa de 5 segundos
    sleep(5)
    # Retorna o elemento selecionado da página.Obs: Usar aspas diferentes.
    campo_pesquisa = driver.find_element(By.XPATH, "//input[@id='cpfInput']")
    sleep(1)
    # Limpa o campo de pesquisa
    campo_pesquisa.clear()
    # Permite escrever algo no campo selecionado.
    campo_pesquisa.send_keys(cpf)
    sleep(1)

    # 3 - Verificar se está em dia ou atrasado.
    # Seleciona o botão da página
    botao_pesquisar = driver.find_element(
        By.XPATH, "//button[@class='btn btn-custom btn-lg btn-block mt-3']")
    sleep(1)
    # Clicar no botão
    botao_pesquisar.click()
    # Dar um tempo pois o resultado demora alguns segundos para aparecer
    sleep(4)

    status = driver.find_element(By.XPATH, "//span[@id='statusLabel']")
    # Cria uma condição de acordo com o resultado do status
    if status.text == 'em dia':
        # 4 - Se estiver "Em dia", pegar a data do pagamento e o método de pagamento
        data_pagamento = driver.find_element(
            By.XPATH, "//p[@id='paymentDate']")
        metodo_pagamento = driver.find_element(
            By.XPATH, "//p[@id='paymentMethod']")

        # Seleciona só o necessário para a consulta
        data_pagamento_limpo = data_pagamento.text.split()[3]
        metodo_pagamento_limpo = metodo_pagamento.text.split()[3]

        planilha_de_fechamento = openpyxl.load_workbook('planilha_fechamento.xlsx')
        pagina_fechamento = planilha_de_fechamento['Sheet1']

        pagina_fechamento.append([nome, valor, cpf, vencimento, 'em dia', data_pagamento_limpo, metodo_pagamento_limpo])

        planilha_de_fechamento.save('planilha_fechamento.xlsx')
    else:
        # 5 - Caso contrário(se estiver atrasado), colocaar o status como pendente
        # explora a planilha de fechamento
        planilha_de_fechamento = openpyxl.load_workbook('planilha_fechamento.xlsx')

        # Seleciona uma aba da planilha
        pagina_fechamento = planilha_de_fechamento['Sheet1']

        pagina_fechamento.append([nome, valor, cpf, vencimento, 'pendente'])

        planilha_de_fechamento.save('planilha_fechamento.xlsx')
